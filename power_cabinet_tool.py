#!/usr/bin/env python3
"""
Power Cabinet Battery Scraper (Vendor-agnostic)

- Input: CSV or XLSX containing site_id + ip (controller address)
- Output: CSV report (sorted by lowest battery %)
- Auth: Credentials loaded from environment variables (no secrets in code)
- Profiles: Supports multiple web-UI "profiles" (no vendor names). Profile A is included.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Optional
from dotenv import load_dotenv
load_dotenv()

import requests
import urllib3
from bs4 import BeautifulSoup

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ---------------------------
# Networking defaults
# ---------------------------
GET_TIMEOUT = (3, 15)    # (connect, read)
POST_TIMEOUT = (3, 35)   # (connect, read)

DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Connection": "close",
}

# ---------------------------
# Credentials (env only)
# ---------------------------

def init_env_file(path: str = ".env"):
    if os.path.exists(path):
        print(f"❌ {path} already exists. Not overwriting.")
        return

    template = (
        "CABINET_USER=Admin\n"
        "CABINET_PASS_1=\n"
        "CABINET_PASS_2=\n"
        "# Add more passwords as needed:\n"
        "# CABINET_PASS_3=\n"
    )

    with open(path, "w", encoding="utf-8") as f:
        f.write(template)

    print(f"✅ Created {path}")
    print("➡️  Edit this file and add your credentials before running the tool.")

def load_credentials(max_passwords: int = 5) -> list[dict]:
    """
    Load credentials from environment variables.

    Required:
      - CABINET_PASS_1 (and optionally CABINET_PASS_2..CABINET_PASS_N)

    Optional:
      - CABINET_USER (defaults to "Admin")
    """
    user = os.getenv("CABINET_USER", "Admin").strip() or "Admin"
    creds: list[dict] = []

    for i in range(1, max_passwords + 1):
        pw = os.getenv(f"CABINET_PASS_{i}", "").strip()
        if pw:
            creds.append({"username": user, "password": pw})

    if not creds:
        raise RuntimeError(
            "No credentials found. Set CABINET_PASS_1 (and optionally CABINET_PASS_2..N). "
            "Optional: CABINET_USER."
        )

    return creds


# ---------------------------
# Input parsing (CSV or XLSX)
# ---------------------------
def read_sites_file(path: str) -> list[dict]:
    """
    Reads either .csv or .xlsx and returns a list of dicts:
      [{"site_id": "...", "ip": "..."}, ...]
    """
    ext = os.path.splitext(path.lower())[1]
    sites: list[dict] = []

    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_l = {(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}

                site_id = row_l.get("site_id") or row_l.get("site") or row_l.get("siteid") or ""
                ip = row_l.get("ip") or row_l.get("controller_ip") or row_l.get("controllerip") or ""

                if ip:
                    sites.append({"site_id": site_id, "ip": ip})

        return sites

    if ext == ".xlsx":
        # Lazy import so CSV users don't need openpyxl installed.
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(path, data_only=True)
        ws = wb.active

        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        headers_norm = [h.strip().lower() for h in headers]

        def col(candidates: list[str]) -> Optional[int]:
            for n in candidates:
                nlow = n.lower()
                if nlow in headers_norm:
                    return headers_norm.index(nlow)
            return None

        site_col = col(["site_id", "site", "siteid"])
        ip_col = col(["ip", "controller_ip", "controllerip"])

        if ip_col is None:
            raise ValueError(
                "Could not find IP column in XLSX. Expected header like: ip / controller_ip."
            )

        for row in ws.iter_rows(min_row=2, values_only=True):
            ip = str(row[ip_col]).strip() if row[ip_col] else ""
            site_id = str(row[site_col]).strip() if site_col is not None and row[site_col] else ""
            if ip:
                sites.append({"site_id": site_id, "ip": ip})

        return sites

    raise ValueError(f"Unsupported file type: {ext} (use .csv or .xlsx)")


# ---------------------------
# Utilities
# ---------------------------
def mask_ip(ip: str) -> str:
    parts = ip.split(".")
    if len(parts) == 4:
        return ".".join(parts[:3] + ["xxx"])
    return ip


def parse_int(s: Optional[str]) -> Optional[int]:
    if not s:
        return None
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else None


def parse_labeled_value(html: str, label_contains: str) -> Optional[str]:
    """
    Find a <td class="label"> containing label_contains, then return the adjacent value cell,
    either from input.value or from cell text.
    """
    soup = BeautifulSoup(html, "html.parser")

    label_td = None
    for td in soup.find_all("td", class_="label"):
        if label_contains.lower() in td.get_text(" ", strip=True).lower():
            label_td = td
            break

    if not label_td:
        return None

    value_td = label_td.find_next_sibling("td") or label_td.find_next("td")
    if not value_td:
        return None

    inp = value_td.find("input")
    if inp and inp.get("value") is not None:
        return inp.get("value")

    return value_td.get_text(" ", strip=True)


def request_with_fallback(session: requests.Session, method: str, ip: str, path: str, **kwargs) -> requests.Response:
    """
    Try HTTPS first (self-signed OK), then fallback to HTTP.
    """
    headers = kwargs.pop("headers", {})
    merged_headers = dict(DEFAULT_HEADERS)
    merged_headers.update(headers)
    kwargs["headers"] = merged_headers

    https_url = f"https://{ip}{path}"
    http_url = f"http://{ip}{path}"

    try:
        return session.request(method, https_url, verify=False, timeout=GET_TIMEOUT, **kwargs)
    except requests.exceptions.RequestException:
        return session.request(method, http_url, timeout=GET_TIMEOUT, **kwargs)


# ---------------------------
# Profiles (vendor-agnostic)
# ---------------------------
@dataclass(frozen=True)
class WebProfile:
    name: str
    login_path: str
    data_path: str
    # detection tokens (any match => candidate)
    detect_any: tuple[str, ...]
    # login form field names
    username_field: str
    password_field: str
    submit_name: str
    submit_value: str
    # labels for scraping
    soc_label: str
    backup_label: str


# Profile A is the one you validated in the field.
# No vendor names here; just a profile identifier.
PROFILE_A = WebProfile(
    name="profile_a",
    login_path="/controller/login",
    data_path="/controller/home",
    detect_any=("inputelement_un", "/controller/"),
    username_field="inputElement_un",
    password_field="inputElement_pw",
    submit_name="formButton_submit",
    submit_value="Login",
    soc_label="State of Charge",
    backup_label="Remaining Backup Time",
)

PROFILES: tuple[WebProfile, ...] = (PROFILE_A,)


def detect_profiles(login_html: str) -> list[WebProfile]:
    """
    Score profiles by how many detect_any tokens appear; return best matches.
    If nothing matches, return all profiles as fallback.
    """
    text = login_html.lower()
    scored: list[tuple[int, WebProfile]] = []

    for p in PROFILES:
        hits = sum(1 for tok in p.detect_any if tok.lower() in text)
        scored.append((hits, p))

    scored.sort(key=lambda t: t[0], reverse=True)
    best = scored[0][0] if scored else 0
    if best == 0:
        return [p for _, p in scored]
    return [p for hits, p in scored if hits == best]


def build_login_payload(profile: WebProfile, login_html: str, username: str, password: str) -> dict:
    soup = BeautifulSoup(login_html, "html.parser")
    payload: dict = {}

    # include hidden inputs if present (future-proof)
    for i in soup.find_all("input", {"type": "hidden"}):
        if i.get("name"):
            payload[i["name"]] = i.get("value", "")

    payload[profile.username_field] = username
    payload[profile.password_field] = password
    payload[profile.submit_name] = profile.submit_value
    return payload


def try_profile(session: requests.Session, ip: str, profile: WebProfile, username: str, password: str) -> dict:
    """
    Attempt login + scrape using a specific profile and credential.
    Returns a result dict with status OK/LOGIN_FAIL/PARSE_FAIL.
    """
    # GET login
    r1 = request_with_fallback(session, "GET", ip, profile.login_path)
    r1.raise_for_status()

    payload = build_login_payload(profile, r1.text, username, password)

    # POST login (prefer HTTPS for cookie behaviors)
    r2 = session.post(
        f"https://{ip}{profile.login_path}",
        data=payload,
        verify=False,
        timeout=POST_TIMEOUT,
        allow_redirects=True,
        headers=DEFAULT_HEADERS,
    )
    r2.raise_for_status()

    # GET data page
    r3 = request_with_fallback(session, "GET", ip, profile.data_path)
    r3.raise_for_status()

    # If still looks like login, assume auth failed
    if profile.username_field.lower() in r3.text.lower() and profile.password_field.lower() in r3.text.lower():
        return {"ip": ip, "status": "LOGIN_FAIL", "profile": profile.name}

    soc_raw = parse_labeled_value(r3.text, profile.soc_label)
    backup_raw = parse_labeled_value(r3.text, profile.backup_label)

    soc = parse_int(soc_raw)
    backup_min = parse_int(backup_raw)

    if soc is None:
        return {"ip": ip, "status": "PARSE_FAIL", "profile": profile.name}

    return {
        "ip": ip,
        "status": "OK",
        "battery_pct": soc,
        "backup_min": backup_min,
        "profile": profile.name,
    }


def fetch_power_cabinet(ip: str, creds: list[dict], debug_dir: str = "debug") -> dict:
    """
    Top-level scrape:
      - detect best profile(s)
      - try each credential
      - return first success
    """
    os.makedirs(debug_dir, exist_ok=True)

    # Use one session per IP for cookie continuity
    s = requests.Session()
    s.headers.update(DEFAULT_HEADERS)

    # Pull login HTML for profile detection
    r = request_with_fallback(s, "GET", ip, PROFILE_A.login_path)
    r.raise_for_status()
    candidates = detect_profiles(r.text)

    last_fail: Optional[dict] = None

    for profile in candidates:
        for c in creds:
            try:
                res = try_profile(s, ip, profile, c["username"], c["password"])
                if res.get("status") == "OK":
                    return res
                last_fail = res
            except Exception as e:
                last_fail = {"ip": ip, "status": "ERROR", "profile": profile.name, "note": str(e)}

    # Save a debug snapshot if we can (helpful for new layouts)
    try:
        safe_ip = ip.replace(".", "_")
        with open(os.path.join(debug_dir, f"login_{safe_ip}.html"), "w", encoding="utf-8") as f:
            f.write(r.text)
    except Exception:
        pass

    return last_fail or {"ip": ip, "status": "FAILED"}


# ---------------------------
# Reporting
# ---------------------------
def write_report_csv(path: str, rows: list[dict]) -> None:
    fieldnames = [
        "timestamp",
        "site_id",
        "ip",
        "status",
        "battery_pct",
        "backup_min",
        "profile",
        "note",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def run_from_file(input_file: str, output_csv: str, mask_ips_in_logs: bool = True) -> str:
    sites = read_sites_file(input_file)
    ts = datetime.now().isoformat(timespec="seconds")
    creds = load_credentials()

    results: list[dict] = []

    for s in sites:
        site_id = s.get("site_id", "")
        ip = s["ip"]
        display_ip = mask_ip(ip) if mask_ips_in_logs else ip

        print(f"\n=== {site_id or '(no site_id)'} | {display_ip} ===")

        try:
            r = fetch_power_cabinet(ip, creds=creds)
            r["timestamp"] = ts
            r["site_id"] = site_id
            r.setdefault("note", "")
            results.append(r)
        except Exception as e:
            results.append({
                "timestamp": ts,
                "site_id": site_id,
                "ip": ip,
                "status": "ERROR",
                "battery_pct": "",
                "backup_min": "",
                "profile": "",
                "note": str(e),
            })

    # sort: lowest battery first (unknowns to bottom)
    def sort_key(x: dict) -> int:
        try:
            return int(x.get("battery_pct"))
        except Exception:
            return 9999

    results.sort(key=sort_key)

    # Auto-timestamp output if file exists (Excel lock friendly)
    if os.path.exists(output_csv):
        base, ext = os.path.splitext(output_csv)
        output_csv = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"

    write_report_csv(output_csv, results)
    print(f"\n✅ Report written to {output_csv}")
    return output_csv


# ---------------------------
# CLI
# ---------------------------
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Power cabinet battery scraper (CSV or XLSX input). Credentials via env vars."
    )
    parser.add_argument("input_file", nargs="?", help="Path to input file (.csv or .xlsx) with columns: site_id, ip")
    parser.add_argument("--out", default="battery_report.csv", help="Output CSV file (default: battery_report.csv)")
    parser.add_argument("--no-mask-ip", action="store_true", help="Do not mask IPs in console output")
    parser.add_argument("--debug-dir", default="debug", help="Directory to save debug HTML (default: debug)")
    parser.add_argument("--init-env", action="store_true", help="Create a .env file template and exit")

    args = parser.parse_args()
    if args.init_env:
        init_env_file()
        return
    
    #otherwise, require input file
    if not args.input_file:
        parser.error("the following arguments are required: input_file")


    # Allow user to choose debug directory
    # (debug files should be .gitignored)
    # We pass it through by temporarily setting env or passing to fetch function in future;
    # for now, run_from_file uses default debug dir inside fetch_power_cabinet.
    _ = args.debug_dir  # reserved for future use; keep CLI stable

    run_from_file(args.input_file, args.out, mask_ips_in_logs=not args.no_mask_ip)


if __name__ == "__main__":
    main()
